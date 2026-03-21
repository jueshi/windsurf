try:
    from google import genai as _genai
    _USE_NEW_GENAI = True
except ImportError:
    import google.generativeai as _genai
    _USE_NEW_GENAI = False
import pandas as pd
import re, os, smtplib
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.text import MIMEText

# ===== 用户配置 =====
stock_list = ["AAPL", "NVDA", "MSFT"]
_api_key = os.getenv("GEMINI_API_KEY")
if _USE_NEW_GENAI:
    _genai_client = _genai.Client(api_key=_api_key)
else:
    _genai.configure(api_key=_api_key)
model_name = "gemini-1.5-flash"
summary_file = "stock_summary_report.xlsx"

# 邮件配置
SMTP_SERVER = "smtp.gmail.com"       # SMTP服务器
SMTP_PORT = 465                   # 端口（SSL一般465）
EMAIL_SENDER = "jueshi@gmail.com"
EMAIL_PASSWORD = "xond wlco mygx abyd"  # 邮箱SMTP授权码
EMAIL_RECEIVER = ["jueshi@gmail.com"]  # 接收方列表，可多个

def analyze_stock(stock_code):
    """调用AI分析股票并返回分数"""
    prompt = """
    请分析 {stock_code} 并仅以JSON返回如下字段：
    - buffett_scores: 长度为8的整数数组，每项0-10
    - buffett_total: 整数，0-100
    - canslim_scores: 长度为7的整数数组，每项0-10
    - canslim_total: 整数，0-100
    要求：
    - 仅输出紧凑JSON，不要有任何额外文字或标点
    - 示例：{{"buffett_scores":[1,2,3,4,5,6,7,8],"buffett_total":70,"canslim_scores":[1,2,3,4,5,6,7],"canslim_total":72}}
    """.format(stock_code=stock_code)
    if _USE_NEW_GENAI:
        resp = _genai_client.models.generate_content(
            model=model_name, contents=prompt,
            config={"temperature": 0}
        )
    else:
        model = _genai.GenerativeModel(model_name)
        resp = model.generate_content(
            prompt,
            generation_config={"temperature": 0}
        )
    text = resp.text

    # 解析JSON输出，带有回退策略
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{[\s\S]*\}", text)
        if m:
            data = json.loads(m.group(0))
        else:
            raise ValueError(f"模型输出无法解析为JSON: {text}")

    buffett_scores = data.get("buffett_scores", [])
    canslim_scores = data.get("canslim_scores", [])
    buffett_total = int(data.get("buffett_total", 0))
    canslim_total = int(data.get("canslim_total", 0))

    # 基本校验
    if not (isinstance(buffett_scores, list) and len(buffett_scores) == 8 and all(isinstance(x, int) for x in buffett_scores)):
        raise ValueError(f"buffett_scores 格式不正确: {buffett_scores}; 原文: {text}")
    if not (isinstance(canslim_scores, list) and len(canslim_scores) == 7 and all(isinstance(x, int) for x in canslim_scores)):
        raise ValueError(f"canslim_scores 格式不正确: {canslim_scores}; 原文: {text}")

    if abs(buffett_total - canslim_total) <= 5 and buffett_total >= 70 and canslim_total >= 70:
        investor_type = "平衡型"
    elif buffett_total > canslim_total:
        investor_type = "价值型"
    else:
        investor_type = "成长型"

    return stock_code, buffett_total, canslim_total, investor_type

def score_to_signal(score):
    if score >= 80: return "高"
    elif score >= 65: return "中"
    else: return "低"

# ===== 主程序：分析并生成Excel =====
summary_data = []
for stock in stock_list:
    print(f"分析 {stock} ...")
    code, buffett_total, canslim_total, inv_type = analyze_stock(stock)
    avg_score = round((buffett_total + canslim_total) / 2, 2)
    summary_data.append({
        "股票代码": code,
        "Buffett综合评分": buffett_total,
        "CANSLIM综合评分": canslim_total,
        "平均分": avg_score,
        "信号等级": score_to_signal(avg_score),
        "投资者类型": inv_type,
        "最后分析日期": datetime.today().strftime("%Y-%m-%d")
    })

df_summary = pd.DataFrame(summary_data)
df_summary["排名"] = df_summary["平均分"].rank(ascending=False, method="min").astype(int)
df_summary = df_summary.sort_values("排名")
df_summary.to_excel(summary_file, index=False)

# ===== Excel信号灯颜色 =====
wb = load_workbook(summary_file)
ws = wb.active
signal_col = None
for i, col in enumerate(ws[1], start=1):
    if col.value == "信号等级":
        signal_col = i
        break
if signal_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=signal_col)
        if cell.value == "高":
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif cell.value == "中":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif cell.value == "低":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
wb.save(summary_file)

# ===== 邮件发送函数 =====
def send_email_with_attachment(subject, body, file_path):
    msg = MIMEMultipart()
    msg["From"] = EMAIL_SENDER
    msg["To"] = ", ".join(EMAIL_RECEIVER)
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain"))

    # 附件
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
    msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        print(f"邮件已发送到 {EMAIL_RECEIVER}")

# ===== 发送汇总表 =====
send_email_with_attachment(
    subject=f"股票优先级报告 - {datetime.today().strftime('%Y-%m-%d')}",
    body="您好，这是最新的股票优先级分析报告，已附Excel文件。\n颜色说明：绿色=优先关注，黄色=可关注，红色=暂缓。",
    file_path=summary_file
)
